# -*- coding: utf-8 -*-
"""Excel 导入导出模块"""
from datetime import datetime, date
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


# ---- 列映射(中文表头 <-> 字段名) ----
MATERIAL_HEADERS = [
    ("物资编码", "material_code"),
    ("名称", "name"),
    ("型号", "model"),
    ("单位", "unit"),
    ("备注", "remark"),
]

PURCHASE_ORDER_HEADERS = [
    ("采购订单号", "order_no"),
    ("行号", "order_line"),
    ("物资编码", "material_code"),
    ("名称", "name"),
    ("型号", "model"),
    ("单位", "unit"),
    ("供应商", "supplier"),
    ("订单数量", "order_quantity"),
    ("订单日期", "order_date"),
    ("交货日期", "delivery_date"),
    ("备注", "remark"),
]

ARRIVAL_HEADERS = [
    ("采购订单号", "purchase_order_no"),
    ("行号", "purchase_order_line"),
    ("物资编码", "material_code"),
    ("名称", "name"),
    ("型号", "model"),
    ("单位", "unit"),
    ("来源", "source"),
    ("供应商", "supplier"),
    ("运单号", "waybill_no"),
    ("到货数量", "arrival_quantity"),
    ("包装", "packaging"),
    ("件数", "package_count"),
    ("重量", "weight"),
    ("管库员", "warehouse_keeper"),
    ("验收完成时间", "acceptance_time"),
    ("上架时间", "shelving_time"),
    ("入库单号", "receipt_number"),
    ("到货登记时间", "arrival_time"),
    ("状态", "status"),
    ("备注", "remark"),
]

INBOUND_REPORT_HEADERS = [
    ("采购订单号", "order_no"),
    ("行号", "order_line"),
    ("物资编码", "material_code"),
    ("名称", "name"),
    ("型号", "model"),
    ("单位", "unit"),
    ("供应商", "supplier"),
    ("订单数量", "order_quantity"),
    ("到货数量", "arrival_quantity"),
    ("已入库数量", "stored_quantity"),
    ("未入库数量", "not_stored_quantity"),
    ("入库状态", "inbound_status"),
]

REFERENCE_REPORT_HEADERS = [
    ("采购订单号", "order_no"),
    ("行号", "order_line"),
    ("物资编码", "material_code"),
    ("名称", "name"),
    ("型号", "model"),
    ("单位", "unit"),
    ("供应商", "supplier"),
    ("订单数量", "order_quantity"),
    ("引用数量", "referenced_quantity"),
    ("未引用数量", "unreferenced_quantity"),
    ("引用状态", "reference_status"),
]

HEADER_ALIASES = {
    "order_no": ["采购订单号", "采购订单", "订单号", "采购单号", "合同号", "单据编号"],
    "order_line": ["行号", "行项目", "订单行号", "项目号", "序号"],
    "material_code": ["物资编码", "物料编码", "物料号", "物资编号", "物料编号", "编码"],
    "name": ["名称", "物资名称", "物料名称", "物料描述", "品名", "产品名称"],
    "model": ["型号", "规格型号", "规格", "型号规格"],
    "unit": ["单位", "计量单位", "采购单位"],
    "supplier": ["供应商", "供应商名称", "供货单位"],
    "order_quantity": ["订单数量", "采购数量", "数量", "订单量", "订货数量"],
    "order_date": ["订单日期", "采购日期", "下单日期", "制单日期"],
    "delivery_date": ["交货日期", "计划交货日期", "到货日期", "需求日期"],
    "purchase_order_no": ["采购订单号", "采购订单", "订单号", "采购单号", "合同号", "单据编号"],
    "purchase_order_line": ["行号", "行项目", "订单行号", "项目号", "序号"],
    "source": ["来源", "到货来源"],
    "waybill_no": ["运单号", "物流单号", "快递单号", "运输单号"],
    "arrival_quantity": ["到货数量", "实到数量", "到货量", "入库数量", "数量"],
    "packaging": ["包装", "包装方式", "包装规格"],
    "package_count": ["件数", "包装件数"],
    "weight": ["重量", "重量(kg)", "毛重"],
    "warehouse_keeper": ["管库员", "库管员", "保管员"],
    "acceptance_time": ["验收完成时间", "验收时间"],
    "shelving_time": ["上架时间"],
    "receipt_number": ["入库单号", "入库单", "入库编号"],
    "remark": ["备注", "说明"],
}


HEADER_FILL = PatternFill("solid", fgColor="305496")
HEADER_FONT = Font(name="微软雅黑", size=11, bold=True, color="FFFFFF")
BODY_FONT = Font(name="微软雅黑", size=10)
ALIGN_CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
ALIGN_LEFT = Alignment(horizontal="left", vertical="center", wrap_text=True)
THIN = Side(style="thin", color="BFBFBF")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


def _style_header(ws, ncols):
    for c in range(1, ncols + 1):
        cell = ws.cell(row=1, column=c)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = ALIGN_CENTER
        cell.border = BORDER
    ws.row_dimensions[1].height = 28


def _autosize(ws, headers, sample_rows=200):
    for idx, (header, _) in enumerate(headers, start=1):
        maxlen = len(str(header))
        for r in range(2, min(ws.max_row, sample_rows) + 1):
            v = ws.cell(row=r, column=idx).value
            if v is None:
                continue
            ln = len(str(v))
            if ln > maxlen:
                maxlen = ln
        ws.column_dimensions[get_column_letter(idx)].width = min(
            max(maxlen + 4, 10), 40
        )


def _fmt_dt(v):
    if v is None or v == "":
        return ""
    if isinstance(v, (datetime, date)):
        if isinstance(v, datetime):
            return v.strftime("%Y-%m-%d %H:%M:%S")
        return v.strftime("%Y-%m-%d")
    return str(v)


def _parse_dt(v):
    """把单元格值解析成 'YYYY-mm-dd HH:MM:SS' 或 None"""
    if v is None or v == "":
        return None
    if isinstance(v, datetime):
        return v.strftime("%Y-%m-%d %H:%M:%S")
    if isinstance(v, date):
        return v.strftime("%Y-%m-%d 00:00:00")
    s = str(v).strip()
    if not s:
        return None
    for fmt in (
        "%Y-%m-%d %H:%M:%S",
        "%Y-%m-%d %H:%M",
        "%Y-%m-%d",
        "%Y/%m/%d %H:%M:%S",
        "%Y/%m/%d %H:%M",
        "%Y/%m/%d",
    ):
        try:
            return datetime.strptime(s, fmt).strftime("%Y-%m-%d %H:%M:%S")
        except ValueError:
            continue
    return s  # 兜底:把原字符串塞进去，让 MySQL 自己判断


def _to_number(v, default=0):
    if v is None or v == "":
        return default
    try:
        return float(v)
    except (TypeError, ValueError):
        try:
            return float(str(v).strip())
        except Exception:
            return default


def _to_int(v, default=0):
    f = _to_number(v, default)
    try:
        return int(f)
    except Exception:
        return default


def _norm_header(value):
    text = str(value or "").strip()
    for ch in ("\n", "\r", "\t", " ", "　", "（", "）", "(", ")", "：", ":"):
        text = text.replace(ch, "")
    return text.lower()


def _cell_text(v):
    if v is None:
        return ""
    if isinstance(v, (datetime, date)):
        return _fmt_dt(v)
    return str(v).strip()


# ====================== 模板生成 ======================

def gen_material_template(path):
    wb = Workbook()
    ws = wb.active
    ws.title = "物资基础"
    headers = [h for h, _ in MATERIAL_HEADERS]
    ws.append(headers)
    sample = [
        ["WZ-0001", "镀锌钢管", "DN50", "米", "示例数据"],
        ["WZ-0002", "六角螺栓", "M12×60", "颗", ""],
    ]
    for row in sample:
        ws.append(row)
    _style_header(ws, len(headers))
    for r in range(2, ws.max_row + 1):
        for c in range(1, len(headers) + 1):
            cell = ws.cell(row=r, column=c)
            cell.font = BODY_FONT
            cell.alignment = ALIGN_LEFT
            cell.border = BORDER
    _autosize(ws, MATERIAL_HEADERS)
    ws.freeze_panes = "A2"
    wb.save(path)


def gen_arrival_template(path):
    wb = Workbook()
    ws = wb.active
    ws.title = "到货登记"
    headers = [h for h, _ in ARRIVAL_HEADERS]
    ws.append(headers)
    sample = [
        ["WZ-0001", "镀锌钢管", "DN50", "米", "示例供应商", 100, 5, 320,
         "", "", "", "", "", "待认领", "示例"],
    ]
    for row in sample:
        ws.append(row)
    _style_header(ws, len(headers))
    for r in range(2, ws.max_row + 1):
        for c in range(1, len(headers) + 1):
            cell = ws.cell(row=r, column=c)
            cell.font = BODY_FONT
            cell.alignment = ALIGN_LEFT
            cell.border = BORDER
    _autosize(ws, ARRIVAL_HEADERS)
    ws.freeze_panes = "A2"
    wb.save(path)


# ====================== 导入 ======================

def _build_header_index(ws, schema):
    """返回 {字段名: 列号(1-based)}"""
    header_row = [c.value for c in ws[1]]
    cn_to_en = {cn: en for cn, en in schema}
    alias_to_field = {}
    for cn, en in schema:
        alias_to_field[_norm_header(cn)] = en
        alias_to_field[_norm_header(en)] = en
    for field, aliases in HEADER_ALIASES.items():
        for alias in aliases:
            alias_to_field[_norm_header(alias)] = field
    idx = {}
    for col_no, value in enumerate(header_row, start=1):
        if value is None:
            continue
        key = str(value).strip()
        if key in cn_to_en:
            idx[cn_to_en[key]] = col_no
        elif key in [en for _, en in schema]:
            idx[key] = col_no
        else:
            field = alias_to_field.get(_norm_header(key))
            if field:
                idx[field] = col_no
    return idx


def import_materials_from_excel(path):
    wb = load_workbook(path, data_only=True)
    ws = wb.active
    idx = _build_header_index(ws, MATERIAL_HEADERS)
    if "material_code" not in idx or "name" not in idx:
        raise ValueError("Excel 缺少必填列「物资编码」或「名称」")
    rows = []
    for r in range(2, ws.max_row + 1):
        code_cell = ws.cell(row=r, column=idx["material_code"]).value
        if code_cell is None or str(code_cell).strip() == "":
            continue
        row = {
            "material_code": str(code_cell).strip(),
            "name": str(ws.cell(row=r, column=idx["name"]).value or "").strip(),
            "model": str(ws.cell(row=r, column=idx.get("model", 0)).value or "").strip()
            if idx.get("model") else "",
            "unit": str(ws.cell(row=r, column=idx.get("unit", 0)).value or "").strip()
            if idx.get("unit") else "",
            "remark": str(ws.cell(row=r, column=idx.get("remark", 0)).value or "").strip()
            if idx.get("remark") else "",
        }
        rows.append(row)
    return rows


def import_purchase_orders_from_excel(path):
    wb = load_workbook(path, data_only=True)
    ws = wb.active
    idx = _build_header_index(ws, PURCHASE_ORDER_HEADERS)
    if "order_no" not in idx or "material_code" not in idx:
        raise ValueError("Excel 缺少必填列「采购订单号」或「物资编码」")
    if "order_quantity" not in idx:
        raise ValueError("Excel 缺少数量列，请确认表头包含「订单数量 / 采购数量 / 数量」")

    rows = []
    for r in range(2, ws.max_row + 1):
        order_no = ws.cell(row=r, column=idx["order_no"]).value
        material_code = ws.cell(row=r, column=idx["material_code"]).value
        if not _cell_text(order_no) or not _cell_text(material_code):
            continue

        def _v(field):
            col = idx.get(field)
            return _cell_text(ws.cell(row=r, column=col).value) if col else ""

        rows.append({
            "order_no": _v("order_no"),
            "order_line": _v("order_line"),
            "material_code": _v("material_code"),
            "name": _v("name"),
            "model": _v("model"),
            "unit": _v("unit"),
            "supplier": _v("supplier"),
            "order_quantity": _to_number(
                ws.cell(row=r, column=idx["order_quantity"]).value
            ),
            "order_date": _v("order_date"),
            "delivery_date": _v("delivery_date"),
            "remark": _v("remark"),
        })
    return rows


def import_arrivals_from_excel(path):
    wb = load_workbook(path, data_only=True)
    ws = wb.active
    idx = _build_header_index(ws, ARRIVAL_HEADERS)
    if "material_code" not in idx:
        raise ValueError("Excel 缺少必填列「物资编码」")
    rows = []
    for r in range(2, ws.max_row + 1):
        code_cell = ws.cell(row=r, column=idx["material_code"]).value
        if code_cell is None or str(code_cell).strip() == "":
            continue

        def _v(field):
            col = idx.get(field)
            if not col:
                return ""
            v = ws.cell(row=r, column=col).value
            return "" if v is None else str(v).strip()

        row = {
            "purchase_order_no": _v("purchase_order_no"),
            "purchase_order_line": _v("purchase_order_line"),
            "material_code": _v("material_code"),
            "name": _v("name"),
            "model": _v("model"),
            "unit": _v("unit"),
            "source": _v("source"),
            "supplier": _v("supplier"),
            "waybill_no": _v("waybill_no"),
            "arrival_quantity": _to_number(
                ws.cell(row=r, column=idx["arrival_quantity"]).value
                if idx.get("arrival_quantity") else 0
            ),
            "packaging": _v("packaging"),
            "package_count": _to_int(
                ws.cell(row=r, column=idx["package_count"]).value
                if idx.get("package_count") else 0
            ),
            "weight": _to_number(
                ws.cell(row=r, column=idx["weight"]).value
                if idx.get("weight") else 0
            ),
            "warehouse_keeper": _v("warehouse_keeper"),
            "acceptance_time": _parse_dt(
                ws.cell(row=r, column=idx["acceptance_time"]).value
                if idx.get("acceptance_time") else None
            ),
            "shelving_time": _parse_dt(
                ws.cell(row=r, column=idx["shelving_time"]).value
                if idx.get("shelving_time") else None
            ),
            "receipt_number": _v("receipt_number"),
            "remark": _v("remark"),
        }
        rows.append(row)
    return rows


# ====================== 导出 ======================

def export_records(path, records, schema, sheet_name="数据"):
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name
    headers = [h for h, _ in schema]
    ws.append(headers)
    for rec in records:
        line = []
        for _, field in schema:
            v = rec.get(field, "")
            if field in ("acceptance_time", "shelving_time", "arrival_time",
                        "created_at", "updated_at"):
                v = _fmt_dt(v)
            elif v is None:
                v = ""
            line.append(v)
        ws.append(line)

    _style_header(ws, len(headers))
    for r in range(2, ws.max_row + 1):
        for c in range(1, len(headers) + 1):
            cell = ws.cell(row=r, column=c)
            cell.font = BODY_FONT
            cell.alignment = ALIGN_LEFT
            cell.border = BORDER

    _autosize(ws, schema)
    ws.freeze_panes = "A2"
    wb.save(path)


def export_materials(path, records):
    export_records(path, records, MATERIAL_HEADERS, "物资基础")


def export_arrivals(path, records):
    export_records(path, records, ARRIVAL_HEADERS, "到货入库进度")


def export_inbound_status(path, records):
    export_records(path, records, INBOUND_REPORT_HEADERS, "已到货物资入库情况")


def export_purchase_order_reference(path, records):
    export_records(path, records, REFERENCE_REPORT_HEADERS, "采购订单引用情况")
